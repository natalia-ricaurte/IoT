import streamlit as st
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.chart import RadarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import io
import json

from utils.constants import METRIC_NAMES_ES, METRIC_NAMES, EXPORT_COLUMN_MAPPING, RECOMMENDED_WEIGHTS
from utils.helpers import to_dict_flat
from weights import validate_manual_weights

class ExcelExporter:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws_summary = self.wb.active
        self.ws_summary.title = "Resumen"
        self.calculation_date = st.session_state.get('global_calculation_date', 
                                                   datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    def _create_summary_header(self):
        """Creates the header for the summary sheet."""
        self.ws_summary['A1'] = "Dashboard de Evaluación de Sostenibilidad - Dispositivos IoT"
        self.ws_summary['A1'].font = Font(bold=True, size=14)
        self.ws_summary['A3'] = f"Fecha y hora del cálculo: {self.calculation_date}"
        self.ws_summary['A4'] = f"Índice de Sostenibilidad Global: {st.session_state.global_result['total_average']:.2f}/10"

    def _create_weights_table(self, ws, start_row, weights):
        """Creates a table of weights in the specified worksheet."""
        ws[f'A{start_row}'] = "Pesos utilizados"
        ws[f'A{start_row}'].font = Font(bold=True)
        ws[f'A{start_row+1}'] = "Métrica"
        ws[f'B{start_row+1}'] = "Peso"
        ws[f'A{start_row+1}'].font = ws[f'B{start_row+1}'].font = Font(bold=True)
        
        for i, k in enumerate(METRIC_NAMES.keys()):
            weight_value = weights.get(k, None)
            if isinstance(weight_value, dict):
                weight_value = list(weight_value.values())[0]
            try:
                weight_value = float(weight_value)
            except Exception:
                weight_value = ''
            ws[f'A{start_row+2+i}'] = METRIC_NAMES_ES[k]
            ws[f'B{start_row+2+i}'] = weight_value
        
        return start_row + 2 + len(METRIC_NAMES)

    def _create_radar_chart(self, ws, start_row, data, title, position, value_col=2, category_col=1):
        """Creates a radar chart in the specified worksheet."""
        chart = RadarChart()
        chart.type = "standard"
        chart.style = 2
        chart.title = title
        data_ref = Reference(ws, min_col=value_col, min_row=start_row, max_row=start_row+len(data)-1)
        cats_ref = Reference(ws, min_col=category_col, min_row=start_row, max_row=start_row+len(data)-1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 10
        ws.add_chart(chart, position)

    def _create_summary_sheet(self):
        """Creates the summary sheet with all its components."""
        self._create_summary_header()
        
        # Global metrics chart
        self.ws_summary['A6'] = "Análisis de Desempeño Ambiental - Métricas Globales"
        self.ws_summary['A6'].font = Font(bold=True, size=12)
        
        # Añadir nota explicativa sobre los colores
        self.ws_summary['A7'] = "Nota: Los colores indican el nivel de desempeño de cada métrica:"
        self.ws_summary['A7'].font = Font(bold=True)
        self.ws_summary['A8'] = "🟢 Verde: Alto desempeño (8-10)"
        self.ws_summary['A8'].font = Font(color="5fba7d")
        self.ws_summary['A9'] = "🟡 Amarillo: Desempeño moderado (5-7.99)"
        self.ws_summary['A9'].font = Font(color="ffd700")
        self.ws_summary['A10'] = "🔴 Rojo: Bajo desempeño (< 5)"
        self.ws_summary['A10'].font = Font(color="ff6b6b")
        
        # Añadir guía de interpretación
        self.ws_summary['A11'] = "Guía de interpretación:"
        self.ws_summary['A11'].font = Font(bold=True)
        self.ws_summary['A12'] = "• Valores cercanos a 10 indican alto desempeño ambiental"
        self.ws_summary['A13'] = "• Valores bajos indican áreas con potencial de mejora"
        
        metrics = list(METRIC_NAMES_ES.values())
        values = [st.session_state.global_result['metrics_average'][k] for k in METRIC_NAMES_ES.keys()]
        
        metrics_row = 15  # Ajustado para dejar espacio a las notas
        
        for i, (metric, value) in enumerate(zip(metrics, values)):
            self.ws_summary[f'A{metrics_row+i}'] = metric
            self.ws_summary[f'B{metrics_row+i}'] = value
            
            # Aplicar color según el valor
            if value >= 8:
                color = "E8F5E9"  # Verde muy suave
            elif value >= 5:
                color = "FFFDE7"  # Amarillo muy suave
            else:
                color = "FFEBEE"  # Rojo muy suave
                
            self.ws_summary[f'B{metrics_row+i}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Añadir métricas destacadas
        highlight_row = metrics_row + len(metrics) + 2
        self.ws_summary[f'A{highlight_row}'] = "Métricas más destacadas:"
        self.ws_summary[f'A{highlight_row}'].font = Font(bold=True)
        
        # Encontrar mejor y peor métrica
        best_idx = values.index(max(values))
        worst_idx = values.index(min(values))
        
        self.ws_summary[f'A{highlight_row+1}'] = f"🏆 Mejor desempeño: {metrics[best_idx]} ({values[best_idx]:.2f})"
        self.ws_summary[f'A{highlight_row+2}'] = f"⚠️ Necesita atención: {metrics[worst_idx]} ({values[worst_idx]:.2f})"
        
        self._create_radar_chart(
            self.ws_summary, 
            metrics_row,  # Start row for normalized metrics
            values,
            "Promedio de Métricas Normalizadas",
            f"G{metrics_row-2}"  # Cambiado de D a G para mover 3 columnas a la derecha
        )

    def _create_devices_sheet(self):
        """Creates the sheet with the list of devices."""
        ws_devices = self.wb.create_sheet("Dispositivos")
        ws_devices['A1'] = "Lista de Dispositivos Evaluados"
        ws_devices['A1'].font = Font(bold=True, size=14)

        internal_columns = [
            "name", "power", "hours", "days", "weight", "life", "renewable_energy", "functionality", "recyclability",
            "B", "Wb", "M", "C", "Wc", "W0", "W"
        ]
        headers = [
            "Nombre del dispositivo", "Potencia (W)", "Horas de uso diario", "Días de uso al año", "Peso (kg)", "Vida útil (años)",
            "Energía renovable (%)", "Funcionalidad (1-10)", "Reciclabilidad (%)", "Baterías vida útil", "Peso batería (g)",
            "Mantenimientos", "Componentes reemplazados", "Peso componente (g)", "Peso nuevo (g)", "Peso final (g)"
        ]
        headers.append("Índice de Sostenibilidad")
        headers.append("Incluido en cálculo global")  # New column

        for i, header in enumerate(headers):
            ws_devices[f'{get_column_letter(i+1)}3'] = header
            ws_devices[f'{get_column_letter(i+1)}3'].font = Font(bold=True)

        # Highlight only the index column
        index_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow
        inclusion_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Very light gray
        
        idx_col = len(internal_columns) + 1
        inclusion_col = idx_col + 1
        
        # Highlight index column
        idx_cell = ws_devices[f'{get_column_letter(idx_col)}3']
        idx_cell.font = Font(bold=True)
        idx_cell.fill = index_fill
        
        # More subtle style for inclusion column
        inclusion_cell = ws_devices[f'{get_column_letter(inclusion_col)}3']
        inclusion_cell.font = Font(bold=True)
        inclusion_cell.fill = inclusion_fill

        for i, device in enumerate(st.session_state.devices):
            # Device data
            for j, col in enumerate(internal_columns):
                value = device.get(col, 'N/A')
                ws_devices[f'{get_column_letter(j+1)}{i+4}'] = value
            
            # Sustainability index (highlighted)
            val_cell = ws_devices[f'{get_column_letter(idx_col)}{i+4}']
            val_cell.value = device['result']['sustainability_index']
            val_cell.fill = index_fill
            val_cell.font = Font(bold=True)
            
            # Global calculation inclusion (subtle style)
            inclusion_cell = ws_devices[f'{get_column_letter(inclusion_col)}{i+4}']
            inclusion_cell.value = "Sí" if st.session_state.selected_devices.get(device['id'], True) else "No"
            inclusion_cell.fill = inclusion_fill
            inclusion_cell.font = Font(bold=False)  # No bold for the value

    def _create_device_detail_sheet(self, device):
        """Creates a detail sheet for a specific device."""
        ws_detail = self.wb.create_sheet(f"Detalle_{device['name'][:20]}")
        ws_detail['A1'] = f"Detalles del Dispositivo: {device['name']}"
        ws_detail['A1'].font = Font(bold=True, size=14)
        ws_detail['D1'] = f"Índice de Sostenibilidad: {device['result']['sustainability_index']:.2f}/10"
        ws_detail['D1'].font = Font(bold=True, size=14)

        # Add global calculation inclusion note
        included = st.session_state.selected_devices.get(device['id'], True)
        ws_detail['A2'] = f"Estado en cálculo global: {'Incluido' if included else 'No incluido'}"
        ws_detail['A2'].font = Font(bold=True, italic=True)
        if not included:
            ws_detail['A2'].font = Font(bold=True, italic=True, color="808080")  # Gris para dispositivos no incluidos

        ws_detail['A3'] = "Datos de Entrada"
        ws_detail['A3'].font = Font(bold=True)
        internal_columns = [
            "name", "power", "hours", "days", "weight", "life", "renewable_energy", "functionality", "recyclability",
            "B", "Wb", "M", "C", "Wc", "W0", "W"
        ]
        headers = [
            "Nombre del dispositivo", "Potencia (W)", "Horas de uso diario", "Días de uso al año", "Peso (kg)", "Vida útil (años)",
            "Energía renovable (%)", "Funcionalidad (1-10)", "Reciclabilidad (%)", "Baterías vida útil", "Peso batería (g)",
            "Mantenimientos", "Componentes reemplazados", "Peso componente (g)", "Peso nuevo (g)", "Peso final (g)"
        ]
        row = 4
        ws_detail[f'A{row}'] = "Campo"
        ws_detail[f'B{row}'] = "Valor"
        ws_detail[f'A{row}'].font = ws_detail[f'B{row}'].font = Font(bold=True)
        for i, (col, desc) in enumerate(zip(internal_columns, headers)):
            value = device.get(col, 'N/A')
            ws_detail[f'A{row+1+i}'] = desc
            ws_detail[f'B{row+1+i}'] = value
        row_after_table = row + 1 + len(internal_columns)

        # Weights configuration
        config_name = self._get_device_configuration_name(device)
        ws_detail[f'D{row_after_table+2}'] = f"Configuración de pesos utilizada: {config_name}"
        ws_detail[f'D{row_after_table+2}'].font = Font(bold=True)

        #Weights table
        self._create_weights_table(ws_detail, row_after_table+2, device.get('used_weights', {}))

        # Normalized metrics and chart
        ws_detail[f'G5'] = "Métricas Normalizadas"
        ws_detail[f'G5'].font = Font(bold=True)
        metrics = []
        values = []
        metrics_row = 6
        for i, (key, value) in enumerate(device['result']['normalized_metrics'].items()):
            ws_detail[f'G{metrics_row+i}'] = METRIC_NAMES_ES[key]
            ws_detail[f'H{metrics_row+i}'] = value
            metrics.append(METRIC_NAMES_ES[key])
            values.append(value)
        self._create_radar_chart(
            ws_detail,
            metrics_row,
            values,
            f"Métricas Normalizadas - {device['name']}",
            "J5",
            value_col=8,      # Column H for values
            category_col=7    # Column G for categories
        )

    def _get_device_configuration_name(self, device):
        """Gets the name of the weight configuration used for a device.
    
        If the configuration name is stored in the snapshot, it is used.
        Otherwise, it attempts to reconstruct the name from the weight values.
        """
        snapshot = device.get('weights_snapshot', {})
        if 'config_name' in snapshot and snapshot['config_name']:
            return snapshot['config_name']
        used_weights = device.get('used_weights', {})
        weight_mode = snapshot.get('mode', None)  # Try to get mode from snapshot, fallback to None
    
        try:
            cleaned_weights = {
                k: float(list(v.values())[0]) if isinstance(v, dict) else float(v)
                for k, v in used_weights.items()
            }
        except Exception:
            return "Desconocido"

        if weight_mode == "Ajuste Manual":
            recommended = RECOMMENDED_WEIGHTS
            if to_dict_flat(cleaned_weights) == to_dict_flat(recommended):
                return "Pesos Recomendados"
        
            for name, config in st.session_state.saved_weights.items():
                if to_dict_flat(config) == to_dict_flat(cleaned_weights):
                    return f"Configuración Manual: {name}"
            return "Pesos Manuales Personalizados"

        elif weight_mode == "Calcular nuevos pesos":
            for config_name, config in st.session_state.ahp_configurations.items():
                if to_dict_flat(config['weights']) == to_dict_flat(cleaned_weights):
                    return f"Configuración Calculada: {config_name}"
            return "Pesos Calculados"
    
        return "Pesos Recomendados"

    def export(self):
        """Exports all data to an Excel file."""
        self._create_summary_sheet()
        self._create_devices_sheet()
        
        # Create detail sheets for each device
        for device in st.session_state.devices:
            self._create_device_detail_sheet(device)
        
        excel_file = BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        return excel_file

def export_results_excel():
    """Exports the results to an Excel file."""
    exporter = ExcelExporter()
    return exporter.export()

def export_devices_list(devices, format='excel'):
    """
    Exports the device list in the specified format (excel, csv, json),
    using template column names for maximum compatibility.
    Returns a buffer (Excel), encoded string (CSV) or encoded string (JSON).
    """
    column_mapping = EXPORT_COLUMN_MAPPING
    internal_columns = list(column_mapping.keys())
    headers = [column_mapping[col] for col in internal_columns]
    data = []
    for device in devices:
        row = [device.get(col, '') for col in internal_columns]
        data.append(row)
    df = pd.DataFrame(data, columns=headers)
    if format == 'excel':
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Dispositivos')
        buffer.seek(0)
        return buffer
    elif format == 'csv':
        csv = df.to_csv(index=False, sep=',', encoding='utf-8')
        return csv.encode('utf-8')
    elif format == 'json':
        data_json = [
            {column_mapping[col]: device.get(col, '') for col in internal_columns}
            for device in devices
        ]
        json_str = json.dumps(data_json, indent=2, ensure_ascii=False)
        return json_str.encode('utf-8')
    else:
        raise ValueError('Formato no soportado: ' + format) 