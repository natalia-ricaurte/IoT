import streamlit as st
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.chart import RadarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import io
import json

from utilidades.constantes import NOMBRES_METRICAS, MAPEO_COLUMNAS_EXPORTACION
from utilidades.auxiliares import to_dict_flat, obtener_valor_dispositivo
from pesos import obtener_pesos_recomendados, validar_pesos_manuales

class ExportadorExcel:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws_resumen = self.wb.active
        self.ws_resumen.title = "Resumen"
        self.fecha_calculo = st.session_state.get('fecha_calculo_global', 
                                                datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    def _obtener_configuracion_pesos(self):
        """Obtiene la configuración de pesos actual y su nombre."""
        if st.session_state.modo_pesos_radio == "Calcular nuevos pesos" and 'pesos_ahp' in st.session_state:
            nombre_config = "Pesos Calculados"
            pesos_global = st.session_state.pesos_ahp
            for nombre_config, config in st.session_state.configuraciones_ahp.items():
                if to_dict_flat(config['pesos']) == to_dict_flat(pesos_global):
                    nombre_config = f"Configuración Calculada: {nombre_config}"
                    break
        elif st.session_state.modo_pesos_radio == "Ajuste Manual":
            nombre_config = "Pesos Manuales Personalizados"
            pesos_manuales = {k: st.session_state[f"peso_manual_{k}"] for k in NOMBRES_METRICAS}
            pesos_global, _ = validar_pesos_manuales(pesos_manuales)
            for nombre_config, config in st.session_state.pesos_guardados.items():
                if to_dict_flat(config) == to_dict_flat(pesos_global):
                    nombre_config = f"Configuración Manual: {nombre_config}"
                    break
        else:
            nombre_config = "Pesos Recomendados"
            pesos_global = obtener_pesos_recomendados()
        
        return nombre_config, pesos_global

    def _crear_encabezado_resumen(self):
        """Crea el encabezado de la hoja de resumen."""
        self.ws_resumen['A1'] = "Dashboard de Evaluación de Sostenibilidad - Dispositivos IoT"
        self.ws_resumen['A1'].font = Font(bold=True, size=14)
        self.ws_resumen['A3'] = f"Fecha y hora del cálculo: {self.fecha_calculo}"
        self.ws_resumen['A4'] = f"Índice de Sostenibilidad Global: {st.session_state.resultado_global['promedio_total']:.2f}/10"
        
        nombre_config, _ = self._obtener_configuracion_pesos()
        self.ws_resumen['A5'] = f"Configuración de pesos utilizada: {nombre_config}"

    def _crear_tabla_pesos(self, ws, fila_inicio, pesos):
        """Crea una tabla de pesos en la hoja especificada."""
        ws[f'A{fila_inicio}'] = "Pesos utilizados"
        ws[f'A{fila_inicio}'].font = Font(bold=True)
        ws[f'A{fila_inicio+1}'] = "Métrica"
        ws[f'B{fila_inicio+1}'] = "Peso"
        ws[f'A{fila_inicio+1}'].font = ws[f'B{fila_inicio+1}'].font = Font(bold=True)
        
        for i, k in enumerate(NOMBRES_METRICAS.keys()):
            valor_peso = pesos.get(k, None)
            if isinstance(valor_peso, dict):
                valor_peso = list(valor_peso.values())[0]
            try:
                valor_peso = float(valor_peso)
            except Exception:
                valor_peso = ''
            ws[f'A{fila_inicio+2+i}'] = NOMBRES_METRICAS[k]
            ws[f'B{fila_inicio+2+i}'] = valor_peso
        
        return fila_inicio + 2 + len(NOMBRES_METRICAS)

    def _crear_grafico_radar(self, ws, fila_inicio, datos, categorias, titulo, posicion, col_valores=2, col_categorias=1):
        """Crea un gráfico radar en la hoja especificada."""
        chart = RadarChart()
        chart.type = "standard"
        chart.style = 2
        chart.title = titulo
        data = Reference(ws, min_col=col_valores, min_row=fila_inicio, max_row=fila_inicio+len(datos)-1)
        cats = Reference(ws, min_col=col_categorias, min_row=fila_inicio, max_row=fila_inicio+len(datos)-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 10
        ws.add_chart(chart, posicion)

    def _crear_hoja_resumen(self):
        """Crea la hoja de resumen con todos sus componentes."""
        self._crear_encabezado_resumen()
        _, pesos_global = self._obtener_configuracion_pesos()
        
        # Tabla de pesos
        fila_actual = self._crear_tabla_pesos(self.ws_resumen, 7, pesos_global)
        
        # Gráfico de métricas globales
        self.ws_resumen[f'A{fila_actual+1}'] = "Gráfico de Métricas Globales"
        self.ws_resumen[f'A{fila_actual+1}'].font = Font(bold=True)
        
        metricas = list(NOMBRES_METRICAS.values())
        valores = [st.session_state.resultado_global['promedio_metricas'][k] for k in NOMBRES_METRICAS.keys()]
        
        fila_metricas = fila_actual + 3
        for i, (metrica, valor) in enumerate(zip(metricas, valores)):
            self.ws_resumen[f'A{fila_metricas+i}'] = metrica
            self.ws_resumen[f'B{fila_metricas+i}'] = valor
        
        self._crear_grafico_radar(
            self.ws_resumen, 
            fila_metricas,  # Fila de inicio de métricas normalizadas
            valores, 
            metricas, 
            "Promedio de Métricas Normalizadas",
            f"D{fila_actual+1}"
        )

    def _formatear_encabezado(self, texto):
        # Convierte descripciones largas en títulos más estilizados
        # Ejemplo: "Potencia eléctrica en vatios (W) del dispositivo cuando está en funcionamiento." -> "Potencia (W)"
        if 'vatios' in texto or '(W)' in texto:
            return 'Potencia (W)'
        if 'horas al día' in texto:
            return 'Horas de uso diario'
        if 'días al año' in texto:
            return 'Días de uso al año'
        if 'kilogramos' in texto or '(kg)' in texto:
            return 'Peso (kg)'
        if 'vida útil' in texto and 'años' in texto:
            return 'Vida útil (años)'
        if 'Porcentaje de energía' in texto:
            return 'Energía renovable (%)'
        if 'funcionalidad' in texto:
            return 'Funcionalidad (1-10)'
        if 'reciclable' in texto or 'reciclabilidad' in texto:
            return 'Reciclabilidad (%)'
        if 'baterías' in texto:
            return 'Baterías vida útil'
        if 'batería' in texto and 'gramos' in texto:
            return 'Peso batería (g)'
        if 'mantenimiento' in texto and 'veces' in texto:
            return 'Mantenimientos'
        if 'componentes reemplazados' in texto:
            return 'Componentes reemplazados'
        if 'componente reemplazado' in texto:
            return 'Peso componente (g)'
        if 'nuevo' in texto and 'gramos' in texto:
            return 'Peso nuevo (g)'
        if 'final' in texto and 'gramos' in texto:
            return 'Peso final (g)'
        if 'nombre' in texto.lower():
            return 'Nombre del dispositivo'
        return texto

    def _crear_hoja_dispositivos(self):
        """Crea la hoja con la lista de dispositivos."""
        ws_dispositivos = self.wb.create_sheet("Dispositivos")
        ws_dispositivos['A1'] = "Lista de Dispositivos Evaluados"
        ws_dispositivos['A1'].font = Font(bold=True, size=14)

        columnas_internas = [
            "nombre", "potencia", "horas", "dias", "peso", "vida", "energia_renovable", "funcionalidad", "reciclabilidad",
            "B", "Wb", "M", "C", "Wc", "W0", "W"
        ]
        headers = [
            "Nombre del dispositivo", "Potencia (W)", "Horas de uso diario", "Días de uso al año", "Peso (kg)", "Vida útil (años)",
            "Energía renovable (%)", "Funcionalidad (1-10)", "Reciclabilidad (%)", "Baterías vida útil", "Peso batería (g)",
            "Mantenimientos", "Componentes reemplazados", "Peso componente (g)", "Peso nuevo (g)", "Peso final (g)"
        ]
        headers.append("Índice de Sostenibilidad")

        for i, header in enumerate(headers):
            ws_dispositivos[f'{get_column_letter(i+1)}3'] = header
            ws_dispositivos[f'{get_column_letter(i+1)}3'].font = Font(bold=True)

        # Resaltar la columna del índice
        fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Amarillo claro
        col_idx = len(columnas_internas) + 1
        cell = ws_dispositivos[f'{get_column_letter(col_idx)}3']
        cell.font = Font(bold=True)
        cell.fill = fill

        for i, dispositivo in enumerate(st.session_state.dispositivos):
            for j, col in enumerate(columnas_internas):
                valor = dispositivo.get(col, 'N/A')
                ws_dispositivos[f'{get_column_letter(j+1)}{i+4}'] = valor
            cell_val = ws_dispositivos[f'{get_column_letter(col_idx)}{i+4}']
            cell_val.value = dispositivo['resultado']['indice_sostenibilidad']
            cell_val.fill = fill
            cell_val.font = Font(bold=True)

    def _crear_hoja_detalle_dispositivo(self, dispositivo):
        """Crea una hoja de detalle para un dispositivo específico."""
        ws_detalle = self.wb.create_sheet(f"Detalle_{dispositivo['nombre'][:20]}")
        ws_detalle['A1'] = f"Detalles del Dispositivo: {dispositivo['nombre']}"
        ws_detalle['A1'].font = Font(bold=True, size=14)
        ws_detalle['D1'] = f"Índice de Sostenibilidad: {dispositivo['resultado']['indice_sostenibilidad']:.2f}/10"
        ws_detalle['D1'].font = Font(bold=True, size=14)

        ws_detalle['A3'] = "Datos de Entrada"
        ws_detalle['A3'].font = Font(bold=True)
        columnas_internas = [
            "nombre", "potencia", "horas", "dias", "peso", "vida", "energia_renovable", "funcionalidad", "reciclabilidad",
            "B", "Wb", "M", "C", "Wc", "W0", "W"
        ]
        headers = [
            "Nombre del dispositivo", "Potencia (W)", "Horas de uso diario", "Días de uso al año", "Peso (kg)", "Vida útil (años)",
            "Energía renovable (%)", "Funcionalidad (1-10)", "Reciclabilidad (%)", "Baterías vida útil", "Peso batería (g)",
            "Mantenimientos", "Componentes reemplazados", "Peso componente (g)", "Peso nuevo (g)", "Peso final (g)"
        ]
        fila = 4
        ws_detalle[f'A{fila}'] = "Campo"
        ws_detalle[f'B{fila}'] = "Valor"
        ws_detalle[f'A{fila}'].font = ws_detalle[f'B{fila}'].font = Font(bold=True)
        for i, (col, desc) in enumerate(zip(columnas_internas, headers)):
            valor = dispositivo.get(col, 'N/A')
            ws_detalle[f'A{fila+1+i}'] = desc
            ws_detalle[f'B{fila+1+i}'] = valor
        fila_despues_tabla = fila + 1 + len(columnas_internas)

        # Configuración de pesos
        nombre_config_disp = self._obtener_nombre_configuracion_dispositivo(dispositivo)
        ws_detalle[f'D{fila_despues_tabla+2}'] = f"Configuración de pesos utilizada: {nombre_config_disp}"
        ws_detalle[f'D{fila_despues_tabla+2}'].font = Font(bold=True)
        
        # Tabla de pesos
        fila_actual = self._crear_tabla_pesos(ws_detalle, fila_despues_tabla+2, dispositivo.get('pesos_utilizados', {}))
        
        # Métricas normalizadas y gráfico
        ws_detalle[f'G5'] = "Métricas Normalizadas"
        ws_detalle[f'G5'].font = Font(bold=True)
        metricas = []
        valores = []
        fila_metricas = 6
        for i, (key, value) in enumerate(dispositivo['resultado']['metricas_normalizadas'].items()):
            ws_detalle[f'G{fila_metricas+i}'] = NOMBRES_METRICAS[key]
            ws_detalle[f'H{fila_metricas+i}'] = value
            metricas.append(NOMBRES_METRICAS[key])
            valores.append(value)
        self._crear_grafico_radar(
            ws_detalle,
            fila_metricas,
            valores,
            metricas,
            f"Métricas Normalizadas - {dispositivo['nombre']}",
            "J5",
            col_valores=8,      # Columna H
            col_categorias=7    # Columna G
        )

    def _obtener_nombre_configuracion_dispositivo(self, dispositivo):
        """Obtiene el nombre de la configuración de pesos utilizada por un dispositivo."""
        pesos_disp = dispositivo.get('pesos_utilizados', {})
        modo_disp = dispositivo.get('snapshot_pesos', {}).get('modo', None)
        
        if modo_disp == "Ajuste Manual":
            pesos_limpios = {k: float(list(v.values())[0]) if isinstance(v, dict) else float(v) 
                           for k, v in pesos_disp.items()}
            pesos_recomendados = obtener_pesos_recomendados()
            if to_dict_flat(pesos_limpios) == to_dict_flat(pesos_recomendados):
                return "Pesos Recomendados"
            
            for nombre, config in st.session_state.pesos_guardados.items():
                if to_dict_flat(config) == to_dict_flat(pesos_limpios):
                    return f"Configuración Manual: {nombre}"
            return "Pesos Manuales Personalizados"
            
        elif modo_disp == "Calcular nuevos pesos":
            pesos_limpios = {k: float(list(v.values())[0]) if isinstance(v, dict) else float(v) 
                           for k, v in pesos_disp.items()}
            for nombre_config, config in st.session_state.configuraciones_ahp.items():
                if to_dict_flat(config['pesos']) == to_dict_flat(pesos_limpios):
                    return f"Configuración Calculada: {nombre_config}"
            return "Pesos Calculados"
        
        return "Pesos Recomendados"

    def exportar(self):
        """Exporta todos los resultados a un archivo Excel."""
        self._crear_hoja_resumen()
        self._crear_hoja_dispositivos()
        
        for dispositivo in st.session_state.dispositivos:
            self._crear_hoja_detalle_dispositivo(dispositivo)
        
        excel_file = BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        return excel_file

def exportar_resultados_excel():
    """Función principal para exportar resultados a Excel."""
    exportador = ExportadorExcel()
    return exportador.exportar()

def exportar_lista_dispositivos(dispositivos, formato='excel'):
    """
    Exporta la lista de dispositivos en el formato especificado (excel, csv, json),
    usando los nombres de columna de la plantilla para máxima compatibilidad.
    Retorna un buffer (Excel), string codificado (CSV) o string codificado (JSON).
    """
    mapeo_columnas = MAPEO_COLUMNAS_EXPORTACION
    columnas_internas = list(mapeo_columnas.keys())
    headers = [mapeo_columnas[col] for col in columnas_internas]
    data = []
    for disp in dispositivos:
        row = [disp.get(col, '') for col in columnas_internas]
        data.append(row)
    df = pd.DataFrame(data, columns=headers)
    if formato == 'excel':
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Dispositivos')
        buffer.seek(0)
        return buffer
    elif formato == 'csv':
        csv = df.to_csv(index=False, sep=',', encoding='utf-8')
        return csv.encode('utf-8')
    elif formato == 'json':
        data_json = [
            {mapeo_columnas[col]: disp.get(col, '') for col in columnas_internas}
            for disp in dispositivos
        ]
        json_str = json.dumps(data_json, indent=2, ensure_ascii=False)
        return json_str.encode('utf-8')
    else:
        raise ValueError('Formato no soportado: ' + formato) 