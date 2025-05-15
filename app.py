import streamlit as st
import numpy as np
import pandas as pd
from streamlit_echarts import st_echarts
import uuid
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.chart import RadarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --- Función auxiliar para comparar pesos ---
def to_dict_flat(p):
    if hasattr(p, 'to_dict'):
        return p.to_dict()
    if isinstance(p, pd.Series):
        return p.to_dict()
    if isinstance(p, dict):
        return p
    return dict(p)

from pesos import (
    obtener_pesos_recomendados,
    validar_pesos_manuales,
    NOMBRES_METRICAS,
    razon_consistencia,
    ahp_attributes
)
from modelo import SostenibilidadIoT

def inicializar_pesos_manuales():
    """Inicializa los pesos manuales con los valores recomendados."""
    if 'pesos_manuales' not in st.session_state:
        pesos_recomendados = obtener_pesos_recomendados()
        st.session_state.pesos_manuales = pesos_recomendados.copy()
    for id in NOMBRES_METRICAS:
        if f"peso_manual_{id}" not in st.session_state:
            st.session_state[f"peso_manual_{id}"] = float(pesos_recomendados[id])

def inicializar_estado():
    """Inicializa todas las variables de estado necesarias."""
    if "dispositivos" not in st.session_state:
        st.session_state.dispositivos = []
    
    if 'matriz_ahp_abierta' not in st.session_state:
        st.session_state.matriz_ahp_abierta = False

    if 'pesos_manuales' not in st.session_state:
        inicializar_pesos_manuales()
    
    if 'pesos_guardados' not in st.session_state:
        st.session_state.pesos_guardados = {}
    
    if 'matriz_comparacion' not in st.session_state:
        metricas = list(NOMBRES_METRICAS.keys())
        n = len(metricas)
        st.session_state.matriz_comparacion = np.ones((n, n))

    # Inicializar modo_pesos_radio si no está en modo edición
    if 'modo_pesos_radio' not in st.session_state and not st.session_state.get('modo_edicion', False):
        st.session_state.modo_pesos_radio = "Pesos Recomendados"

    if 'configuraciones_ahp' not in st.session_state:
        st.session_state.configuraciones_ahp = {}

def reiniciar_estado():
    """Reinicia el estado de la aplicación a sus valores iniciales."""
    st.session_state.dispositivos = []
    st.session_state.matriz_ahp_abierta = False
    inicializar_pesos_manuales()
    st.session_state.pesos_guardados = {}
    metricas = list(NOMBRES_METRICAS.keys())
    n = len(metricas)
    st.session_state.matriz_comparacion = np.ones((n, n))
    st.session_state.modo_pesos_radio = "Pesos Recomendados"
    # Eliminar variables adicionales
    for var in [
        'modo_pesos_guardado', 'resultado_global', 'pesos_ahp', 'ahp_resultados',
        'mostrar_tabla_pesos_ahp', 'modo_edicion', 'dispositivo_editando', 'carga_edicion_realizada'
    ]:
        if var in st.session_state:
            del st.session_state[var]

def radar_chart(metricas, titulo, key):
    etiquetas = {
        'CE': 'Cons. Energía', 
        'HC': 'Huella CO₂', 
        'EW': 'E-waste',
        'ER': 'Energía Renov.', 
        'EE': 'Eficiencia', 
        'DP': 'Durabilidad',
        'RC': 'Reciclabilidad', 
        'IM': 'Mantenimiento'
    }
    
    labels = [etiquetas[m] for m in metricas.keys()]
    valores = list(metricas.values())

    options = {
        "backgroundColor": "#111111",
        "title": {
            "text": titulo,
            "left": "center",
            "top": "5%",
            "textStyle": {
                "color": "#ffffff",
                "fontSize": 20,
            },
        },
        "tooltip": {
            "trigger": "item"
        },
        "radar": {
            "indicator": [{"name": label, "max": 10} for label in labels],
            "radius": "60%",
            "center": ["50%", "55%"],
            "splitNumber": 5,
            "axisLine": {
                "lineStyle": {"color": "#3498db"}
            },
            "splitLine": {
                "lineStyle": {"color": "#444444"},
                "show": True
            },
            "splitArea": {
                "areaStyle": {"color": ["#2e2e3e", "#1e1e2f"]}
            }
        },
        "series": [
            {
                "name": titulo,
                "type": "radar",
                "data": [
                    {
                        "value": valores,
                        "name": titulo,
                        "itemStyle": {
                            "color": "#3498db"
                        },
                        "lineStyle": {
                            "color": "#3498db"
                        },
                        "areaStyle": {
                            "color": "#3498db",
                            "opacity": 0.3
                        }
                    }
                ]
            }
        ]
    }

    # Renderizar con ECharts usando el key único
    st_echarts(options=options, height="500px", key=key)


def calcular_pesos_ahp(metricas):
    df_criterios = pd.DataFrame(
        st.session_state.matriz_comparacion,
        index=metricas,
        columns=metricas
    )
    pesos = ahp_attributes(df_criterios)
    ic, rc = razon_consistencia(pesos, df_criterios, verbose=False)
    return pesos, ic, rc

def mostrar_resultados_ahp(pesos, rc):
    """Muestra los resultados del cálculo de pesos por Matriz de Comparación por Pares."""
    st.success("Pesos calculados mediante la Matriz de Comparación por Pares:")
    # Convertir los pesos a un formato limpio
    pesos_limpios = {}
    for k, v in pesos.items():
        if isinstance(v, dict):
            val = list(v.values())[0]
        elif isinstance(v, pd.Series):
            val = v.iloc[0]
        else:
            val = v
        pesos_limpios[k] = float(val)
    # Crear DataFrame con los pesos
    metricas_list = list(pesos_limpios.keys())
    pesos_list = [pesos_limpios[k] for k in metricas_list]
    nombres_list = [NOMBRES_METRICAS[k] for k in metricas_list]
    df_pesos = pd.DataFrame({
        'Métrica': nombres_list,
        'Peso': pesos_list
    })
    # Agregar columna de importancia relativa con criterios equilibrados
    df_pesos['Importancia'] = df_pesos['Peso'].apply(
        lambda x: '🔴 Alta' if x >= 0.20 else '🟡 Media' if 0.10 < x < 0.20 else '🟢 Baja'
    )
    st.dataframe(df_pesos.style.format({'Peso': '{:.3f}'}), use_container_width=True)
    # Mostrar razón de consistencia y advertencias
    if rc < 0.1:
        st.success(f"✅ Razón de Consistencia: {rc:.3f} (La matriz es consistente)")
    else:
        st.warning(f"⚠️ Razón de Consistencia: {rc:.3f} (La matriz NO es consistente)")
        st.info("""
        **Sugerencias para mejorar la consistencia:**
        1. Revise las comparaciones más extremas
        2. Asegúrese de que sus comparaciones sean transitivas
        3. Si A > B y B > C, entonces A debería ser más importante que C
        """)

def mostrar_matriz_ahp():
    st.title("Matriz de Comparación por Pares")
    with st.expander("ℹ️ Guía de la Matriz de Comparación por Pares"):
        st.markdown("""
        ### Guía de la Matriz de Comparación por Pares
        
        Esta matriz le permite comparar la importancia relativa de cada par de métricas utilizando la escala de Saaty.
        
        **Escala de Comparación:**
        - 1: Las métricas son igualmente importantes
        - 3: La métrica de la fila es moderadamente más importante
        - 5: La métrica de la fila es fuertemente más importante
        - 7: La métrica de la fila es muy fuertemente más importante
        - 9: La métrica de la fila es extremadamente más importante
        
        **Valores Decimales (Recíprocos):**
        Cuando una métrica es menos importante, use los siguientes valores decimales:
        | Comparación | Valor Decimal |
        |-------------|---------------|
        | 1/2         | 0.50         |
        | 1/3         | 0.33         |
        | 1/4         | 0.25         |
        | 1/5         | 0.20         |
        | 1/6         | 0.17         |
        | 1/7         | 0.14         |
        | 1/8         | 0.13         |
        | 1/9         | 0.11         |
        
        **Ejemplos:**
        - Si el Consumo de Energía es moderadamente más importante que la Huella de Carbono, ingrese 3
        - Si la Huella de Carbono es fuertemente más importante que el E-waste, ingrese 5
        - Si el E-waste es moderadamente menos importante que la Huella de Carbono, ingrese 0.33 (equivalente a 1/3)
        
        **Consejo:** Comience comparando las métricas más importantes entre sí.
        """)
    st.info("Edita solo la mitad superior de la tabla. El resto se calcula automáticamente.")
    metricas = list(NOMBRES_METRICAS.keys())
    nombres = [NOMBRES_METRICAS[m] for m in metricas]
    n = len(metricas)
    if 'matriz_comparacion' not in st.session_state:
        st.session_state.matriz_comparacion = np.ones((n, n))
    # Encabezados de columna
    cols = st.columns(n+1)
    cols[0].write("")
    for j in range(n):
        cols[j+1].markdown(f"**{nombres[j]}**")
    # Filas de la matriz
    for i in range(n):
        row = st.columns(n+1)
        row[0].markdown(f"**{nombres[i]}**")
        for j in range(n):
            if i == j:
                row[j+1].markdown('<span style="font-weight:bold;">1.00</span>', unsafe_allow_html=True)
                st.session_state.matriz_comparacion[i, j] = 1.0
            elif i < j:
                valor = row[j+1].number_input(
                    f"Comparación {nombres[i]} vs {nombres[j]}",
                    min_value=0.11,
                    max_value=9.0,
                    value=float(st.session_state.matriz_comparacion[i, j]) if st.session_state.matriz_comparacion[i, j] != 1.0 else 1.0,
                    step=0.01,
                    key=f"matriz_{i}_{j}",
                    label_visibility="collapsed"
                )
                st.session_state.matriz_comparacion[i, j] = valor
                st.session_state.matriz_comparacion[j, i] = 1/valor
            else:
                row[j+1].write(f"{st.session_state.matriz_comparacion[i, j]:.2f}")
    st.markdown("---")
    # Botones de acción
    col_calc, col_save, col_space, col_cancel, col_reset = st.columns([1, 1, 2, 1, 1])
    with col_calc:
        if st.button("Calcular pesos"):
            df_criterios = pd.DataFrame(
                st.session_state.matriz_comparacion,
                index=metricas,
                columns=metricas
            )
            pesos = ahp_attributes(df_criterios)
            ic, rc = razon_consistencia(pesos, df_criterios, verbose=False)
            st.session_state.ahp_resultados = {
                'pesos': pesos,
                'ic': ic,
                'rc': rc
            }
            st.rerun()
    with col_save:
        if st.button("Guardar y salir"):
            if 'ahp_resultados' in st.session_state:
                pesos = st.session_state.ahp_resultados['pesos']
                if hasattr(pesos, 'to_dict'):
                    pesos = pesos.to_dict()
                st.session_state.pesos_ahp = pesos
                # Guardar los resultados para mostrar en la pantalla principal
                st.session_state.mostrar_tabla_pesos_ahp = {
                    'pesos': pesos,
                    'rc': st.session_state.ahp_resultados['rc']
                }
            st.session_state.matriz_ahp_abierta = False
            if 'modo_pesos_guardado' in st.session_state:
                st.session_state.modo_pesos_radio = st.session_state.modo_pesos_guardado
            st.rerun()
    with col_cancel:
        if st.button("Cancelar"):
            if 'ahp_resultados' in st.session_state:
                pesos = st.session_state.ahp_resultados['pesos']
                if hasattr(pesos, 'to_dict'):
                    pesos = pesos.to_dict()
                # Guardar los resultados para mostrar en la pantalla principal
                st.session_state.mostrar_tabla_pesos_ahp = {
                    'pesos': pesos,
                    'rc': st.session_state.ahp_resultados['rc']
                }
            st.session_state.matriz_ahp_abierta = False
            if 'modo_pesos_guardado' in st.session_state:
                st.session_state.modo_pesos_radio = st.session_state.modo_pesos_guardado
            st.rerun()
    with col_reset:
        if st.button("Reiniciar matriz", help="Reinicia la matriz de comparación a valores iniciales (identidad)", key="btn_reiniciar_matriz"):
            st.session_state.matriz_comparacion = np.ones((n, n))
            # Limpiar los pesos AHP y resultados anteriores
            if 'pesos_ahp' in st.session_state:
                del st.session_state.pesos_ahp
            if 'ahp_resultados' in st.session_state:
                del st.session_state.ahp_resultados
            if 'mostrar_tabla_pesos_ahp' in st.session_state:
                del st.session_state.mostrar_tabla_pesos_ahp
            st.warning("¡La matriz de comparación ha sido reiniciada a valores iniciales!")
            st.rerun()
    # Mostrar resultados si existen
    if st.session_state.get('ahp_resultados'):
        mostrar_resultados_ahp(st.session_state.ahp_resultados['pesos'], st.session_state.ahp_resultados['rc'])
    st.stop()

def exportar_resultados_excel():
    """Exporta todos los resultados a un archivo Excel con gráficos."""
    # Asegurar que la fecha esté inicializada
    if 'fecha_calculo_global' not in st.session_state:
        st.session_state.fecha_calculo_global = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen['A1'] = "Dashboard de Evaluación de Sostenibilidad - Dispositivos IoT"
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['A3'] = f"Fecha y hora del cálculo: {st.session_state.fecha_calculo_global}"
    ws_resumen['A4'] = f"Índice de Sostenibilidad Global: {st.session_state.resultado_global['promedio_total']:.2f}/10"
    # Obtener nombre de la configuración de pesos y los pesos globales
    if st.session_state.modo_pesos_radio == "Calcular nuevos pesos" and 'pesos_ahp' in st.session_state:
        nombre_config = "Pesos Calculados"
        pesos_global = st.session_state.pesos_ahp
        for nombre_config, config in st.session_state.configuraciones_ahp.items():
            if to_dict_flat(config['pesos']) == to_dict_flat(pesos_global):
                nombre_config = f"Configuración Calculada: {nombre_config}"
                break
    elif st.session_state.modo_pesos_radio == "Ajuste Manual":
        nombre_config = "Pesos Manuales Personalizados"
        pesos_manuales = {k: st.session_state[f"peso_manual_{k}"] for k in NOMBRES_METRICAS}
        pesos_global, _ = validar_pesos_manuales(pesos_manuales)
        for nombre_config, config in st.session_state.pesos_guardados.items():
            if to_dict_flat(config) == to_dict_flat(pesos_global):
                nombre_config = f"Configuración Manual: {nombre_config}"
                break
    else:
        nombre_config = "Pesos Recomendados"
        pesos_global = obtener_pesos_recomendados()
    ws_resumen['A5'] = f"Configuración de pesos utilizada: {nombre_config}"
    # --- Tabla de pesos utilizados ---
    ws_resumen['A7'] = "Pesos utilizados para el cálculo global"
    ws_resumen['A7'].font = Font(bold=True)
    ws_resumen['A8'] = "Métrica"
    ws_resumen['B8'] = "Peso"
    ws_resumen['A8'].font = ws_resumen['B8'].font = Font(bold=True)
    for i, k in enumerate(NOMBRES_METRICAS.keys()):
        valor_peso = pesos_global[k]
        if isinstance(valor_peso, dict):
            valor_peso = list(valor_peso.values())[0]
        try:
            valor_peso = float(valor_peso)
        except Exception:
            valor_peso = ''
        ws_resumen[f'A{9+i}'] = NOMBRES_METRICAS[k]
        ws_resumen[f'B{9+i}'] = valor_peso
    fila_grafico = 9 + len(NOMBRES_METRICAS) + 1
    ws_resumen[f'A{fila_grafico}'] = "Gráfico de Métricas Globales"
    ws_resumen[f'A{fila_grafico}'].font = Font(bold=True)
    # Preparar datos para el gráfico
    metricas = list(NOMBRES_METRICAS.values())
    valores = [st.session_state.resultado_global['promedio_metricas'][k] for k in NOMBRES_METRICAS.keys()]
    for i, (metrica, valor) in enumerate(zip(metricas, valores)):
        ws_resumen[f'A{fila_grafico+2+i}'] = metrica
        ws_resumen[f'B{fila_grafico+2+i}'] = valor
    chart = RadarChart()
    chart.type = "standard"
    chart.style = 2
    chart.title = "Promedio de Métricas Normalizadas"
    data = Reference(ws_resumen, min_col=2, min_row=fila_grafico+2, max_row=fila_grafico+1+len(metricas))
    cats = Reference(ws_resumen, min_col=1, min_row=fila_grafico+2, max_row=fila_grafico+1+len(metricas))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 10
    ws_resumen.add_chart(chart, f"D{fila_grafico}")
    # Hoja de Dispositivos
    ws_dispositivos = wb.create_sheet("Dispositivos")
    ws_dispositivos['A1'] = "Lista de Dispositivos Evaluados"
    ws_dispositivos['A1'].font = Font(bold=True, size=14)
    headers = ['Nombre', 'Índice de Sostenibilidad']
    for i, header in enumerate(headers):
        ws_dispositivos[f'{get_column_letter(i+1)}3'] = header
        ws_dispositivos[f'{get_column_letter(i+1)}3'].font = Font(bold=True)
    for i, dispositivo in enumerate(st.session_state.dispositivos):
        ws_dispositivos[f'A{i+4}'] = dispositivo['nombre']
        ws_dispositivos[f'B{i+4}'] = dispositivo['resultado']['indice_sostenibilidad']
    # Hoja de Detalles por Dispositivo
    for dispositivo in st.session_state.dispositivos:
        ws_detalle = wb.create_sheet(f"Detalle_{dispositivo['nombre'][:20]}")
        ws_detalle['A1'] = f"Detalles del Dispositivo: {dispositivo['nombre']}"
        ws_detalle['A1'].font = Font(bold=True, size=14)
        ws_detalle['A3'] = f"Índice de Sostenibilidad: {dispositivo['resultado']['indice_sostenibilidad']:.2f}/10"
        # Datos de entrada
        ws_detalle['A5'] = "Datos de Entrada"
        ws_detalle['A5'].font = Font(bold=True)
        datos_entrada = {
            'Potencia (W)': dispositivo['potencia'],
            'Horas uso diario': dispositivo['horas'],
            'Días uso/año': dispositivo['dias'],
            'Peso (kg)': dispositivo['peso'],
            'Vida útil (años)': dispositivo['vida'],
            'Energía renovable (%)': dispositivo['energia_renovable'],
            'Funcionalidad': dispositivo['funcionalidad'],
            'Reciclabilidad (%)': dispositivo['reciclabilidad']
        }
        for i, (key, value) in enumerate(datos_entrada.items()):
            ws_detalle[f'A{i+6}'] = key
            ws_detalle[f'B{i+6}'] = value
        # --- Nombre de la configuración de pesos utilizada por dispositivo ---
        # Determinar el nombre de la configuración igual que en el dashboard
        pesos_disp = dispositivo.get('pesos_utilizados', {})
        nombre_config_disp = "Pesos Recomendados"
        modo_disp = dispositivo.get('snapshot_pesos', {}).get('modo', None)
        if modo_disp == "Ajuste Manual":
            pesos_limpios = {k: float(list(v.values())[0]) if isinstance(v, dict) else float(v) for k, v in pesos_disp.items()}
            pesos_recomendados = obtener_pesos_recomendados()
            if to_dict_flat(pesos_limpios) == to_dict_flat(pesos_recomendados):
                nombre_config_disp = "Pesos Recomendados"
            else:
                nombre_config_default = "Pesos Manuales Personalizados"  # valor por defecto
                nombre_config_encontrada = None
                for nombre, config in st.session_state.pesos_guardados.items():
                    if to_dict_flat(config) == to_dict_flat(pesos_limpios):
                        nombre_config_encontrada = f"Configuración Manual: {nombre}"
                        break
                nombre_config_disp = nombre_config_encontrada if nombre_config_encontrada else nombre_config_default
        elif modo_disp == "Calcular nuevos pesos":
            pesos_limpios = {k: float(list(v.values())[0]) if isinstance(v, dict) else float(v) for k, v in pesos_disp.items()}
            for nombre_config, config in st.session_state.configuraciones_ahp.items():
                if to_dict_flat(config['pesos']) == to_dict_flat(pesos_limpios):
                    nombre_config_disp = f"Configuración Calculada: {nombre_config}"
                    break
            else:
                nombre_config_disp = "Pesos Calculados"
        ws_detalle['D4'] = f"Configuración de pesos utilizada: {nombre_config_disp}"
        ws_detalle['D4'].font = Font(bold=True)
        # --- Tabla de pesos utilizados por dispositivo ---
        ws_detalle['D5'] = "Pesos utilizados"
        ws_detalle['D5'].font = Font(bold=True)
        ws_detalle['D6'] = "Métrica"
        ws_detalle['E6'] = "Peso"
        ws_detalle['D6'].font = ws_detalle['E6'].font = Font(bold=True)
        for j, k in enumerate(NOMBRES_METRICAS.keys()):
            valor_peso = pesos_disp.get(k, None)
            if isinstance(valor_peso, dict):
                valor_peso = list(valor_peso.values())[0]
            try:
                valor_peso = float(valor_peso)
            except Exception:
                valor_peso = ''
            ws_detalle[f'D{7+j}'] = NOMBRES_METRICAS[k]
            ws_detalle[f'E{7+j}'] = valor_peso
        # Métricas normalizadas
        base_radar = 7 + len(NOMBRES_METRICAS) + 1
        ws_detalle[f'G5'] = "Métricas Normalizadas"
        ws_detalle[f'G5'].font = Font(bold=True)
        for i, (key, value) in enumerate(dispositivo['resultado']['metricas_normalizadas'].items()):
            ws_detalle[f'G{6+i}'] = NOMBRES_METRICAS[key]
            ws_detalle[f'H{6+i}'] = value
        # Crear gráfico radar para el dispositivo
        chart = RadarChart()
        chart.type = "standard"
        chart.style = 2
        chart.title = f"Métricas Normalizadas - {dispositivo['nombre']}"
        data = Reference(ws_detalle, min_col=8, min_row=6, max_row=6+len(NOMBRES_METRICAS)-1)
        cats = Reference(ws_detalle, min_col=7, min_row=6, max_row=6+len(NOMBRES_METRICAS)-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 10
        ws_detalle.add_chart(chart, "J5")
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

# --- INICIALIZACIÓN DE LA APLICACIÓN ---
st.set_page_config(page_title="Dashboard Sostenibilidad IoT", layout="wide")

# Inicializar variables básicas
if 'dispositivos' not in st.session_state:
    st.session_state.dispositivos = []

if 'matriz_ahp_abierta' not in st.session_state:
    st.session_state.matriz_ahp_abierta = False

if 'pesos_guardados' not in st.session_state:
    st.session_state.pesos_guardados = {}

if 'matriz_comparacion' not in st.session_state:
    metricas = list(NOMBRES_METRICAS.keys())
    n = len(metricas)
    st.session_state.matriz_comparacion = np.ones((n, n))

if 'modo_pesos_radio' not in st.session_state:
    st.session_state.modo_pesos_radio = "Pesos Recomendados"

if 'configuraciones_ahp' not in st.session_state:
    st.session_state.configuraciones_ahp = {}

# Inicializar pesos manuales si no existen
if 'pesos_manuales' not in st.session_state:
    inicializar_pesos_manuales()

# --- CONTROL DE NAVEGACIÓN ---
if st.session_state.matriz_ahp_abierta:
    mostrar_matriz_ahp()
    st.stop()

# --- INTERFAZ PRINCIPAL ---
st.title("Dashboard de Evaluación de Sostenibilidad - Dispositivos IoT")

# Botón para reiniciar
if st.button("Reiniciar"):
    reiniciar_estado()
    st.rerun()

st.markdown("## Descripción de Métricas")
with st.expander("Ver métricas clave del modelo"):
    st.markdown("""
    **Métricas de sostenibilidad evaluadas:**
    1. **CE - Consumo de Energía:** kWh anuales usados por el dispositivo.
    2. **HC - Huella de Carbono:** kg de CO₂eq emitidos.
    3. **EW - E-waste:** kg de residuos electrónicos generados por año.
    4. **ER - Energía Renovable:** Porcentaje de energía limpia usada.
    5. **EE - Eficiencia Energética:** Relación funcionalidad / consumo.
    6. **DP - Durabilidad del Producto:** Vida útil esperada.
    7. **RC - Reciclabilidad:** Porcentaje de materiales reciclables.
    8. **IM - Mantenimiento:** Impacto de baterías, reemplazos y desgaste.
    """)

with st.expander("Guía rápida de uso del dashboard"):
    st.markdown("""
1. **Define los pesos de las métricas**
   - Selecciona el método de asignación de pesos en la columna derecha antes de ingresar dispositivos.
   - Puedes usar los pesos recomendados, ajustarlos manualmente o calcular nuevos pesos mediante comparación por pares.
   - **Pesos Recomendados:** Basados en análisis y alineación con ODS.
   - **Ajuste Manual:** Permite personalizar los pesos y guardar configuraciones personalizadas.
   - **Pesos Calculados:** Utiliza la matriz de comparación por pares y permite guardar diferentes configuraciones.
   - **Nota:** Los pesos activos al momento de añadir un dispositivo serán los que se usen para su cálculo. El nombre de la configuración utilizada se guarda y se muestra en los resultados y exportaciones.

2. **Ingresa las características de tus dispositivos IoT**
   - Completa el formulario y pulsa **'Añadir dispositivo'** para guardar cada uno.
   - Al añadir un nuevo dispositivo, los resultados globales previos se eliminan automáticamente. Deberás recalcular el índice global para ver los resultados actualizados.

3. **Gestiona tu lista de dispositivos**
   - Puedes ver los detalles completos de cada dispositivo pulsando **'Mostrar detalles'**.
   - Dentro de los detalles, consulta los datos de entrada y los pesos utilizados para ese dispositivo, junto con el nombre de la configuración de pesos aplicada.
   - Para eliminar un dispositivo, marca la casilla **'Eliminar dispositivo'** y confirma la acción con el botón correspondiente. Al eliminar cualquier dispositivo, los resultados globales se eliminan y deberás recalcular.

4. **Calcula y analiza los resultados**
   - Pulsa **'Calcular Índice de Sostenibilidad'** para ver los resultados individuales y globales.
   - El índice global y los detalles del sistema solo reflejan los dispositivos actualmente en la lista.

5. **Consulta los detalles del sistema**
   - En la sección de resultados globales, expande **'Detalles del sistema'** para ver:
     - Cantidad total de dispositivos evaluados.
     - Desviación estándar de los índices individuales.
     - Fecha y hora del cálculo global.
     - Nota sobre comparabilidad si los dispositivos fueron evaluados con diferentes pesos.
     - Pesos utilizados para el cálculo global y nombre de la configuración aplicada.
     - Lista de dispositivos incluidos y su índice individual.

6. **Exporta los resultados completos a Excel**
   - Tras calcular el índice global, utiliza el botón **'Descargar Resultados Completos'** para exportar toda la información a un archivo Excel profesional.
   - El archivo incluye:
     - Resumen general con índice global, fecha, configuración de pesos y gráfico radar.
     - Tabla de pesos utilizados para el cálculo global.
     - Lista de dispositivos y sus índices.
     - Hojas de detalle para cada dispositivo, con datos de entrada, nombre de la configuración de pesos utilizada, tabla de pesos y gráfico radar individual.

---

**Consejos y advertencias:**
- Si cambias los pesos o la lista de dispositivos, recuerda recalcular el índice global para obtener resultados actualizados.
- Si los dispositivos fueron evaluados con diferentes configuraciones de pesos, los índices individuales pueden no ser directamente comparables.
- El dashboard elimina automáticamente los resultados globales al añadir o eliminar dispositivos para evitar mostrar información desactualizada.
- Puedes guardar y cargar diferentes configuraciones de pesos tanto para el ajuste manual como para los pesos calculados mediante comparación por pares.
- El nombre de la configuración de pesos utilizada se guarda y se muestra en todos los resultados y exportaciones para máxima trazabilidad.
""")

# Definir claves y valores por defecto para el formulario
form_keys = {
    'nombre': ("Sensor de temperatura", "Nombre descriptivo del dispositivo IoT."),
    'potencia': (2.0, "Potencia eléctrica en vatios (W) del dispositivo cuando está en funcionamiento."),
    'horas': (24.0, "Cantidad de horas al día que el dispositivo está en uso."),
    'dias': (365, "Número de días al año que el dispositivo opera."),
    'peso': (0.1, "Peso total del dispositivo en kilogramos."),
    'vida': (5, "Duración esperada del dispositivo antes de desecharse o reemplazarse."),
    'energia_renovable': (30, "Porcentaje de energía que proviene de fuentes renovables."),
    'funcionalidad': (8, "Nivel de funcionalidad y utilidad que ofrece el dispositivo."),
    'reciclabilidad': (65, "Porcentaje del dispositivo que puede reciclarse al finalizar su vida útil."),
    'B': (2, "Cantidad de baterías necesarias durante toda la vida útil del dispositivo."),
    'Wb': (50, "Peso de cada batería en gramos."),
    'M': (1, "Número de veces que el dispositivo requiere mantenimiento."),
    'C': (2, "Número de componentes reemplazados en mantenimientos."),
    'Wc': (20, "Peso promedio de cada componente reemplazado en gramos."),
    'W0': (200, "Peso total del dispositivo cuando es nuevo."),
    'W': (180, "Peso final del dispositivo después del uso.")
}
# Inicializar en session_state si no existen
for k, (default, _) in form_keys.items():
    if f"form_{k}" not in st.session_state:
        st.session_state[f"form_{k}"] = default

col1, col2 = st.columns([2, 1])

with col1:
    submitted = False  # Inicializar para evitar NameError
    with st.form("formulario_datos"):
        st.subheader("Datos del dispositivo IoT")
        colA, colB = st.columns(2)
        with colA:
            nombre = st.text_input("Nombre del dispositivo", value=st.session_state["form_nombre"], help="Nombre descriptivo del dispositivo IoT.")
            potencia = st.number_input("Potencia (W)", value=st.session_state["form_potencia"], help="Potencia eléctrica en vatios (W) del dispositivo cuando está en funcionamiento.")
            horas = st.number_input("Horas uso diario", value=st.session_state["form_horas"], help="Cantidad de horas al día que el dispositivo está en uso.")
            dias = st.number_input("Días uso/año", value=st.session_state["form_dias"], help="Número de días al año que el dispositivo opera.")
            peso = st.number_input("Peso dispositivo (kg)", value=st.session_state["form_peso"], help="Peso total del dispositivo en kilogramos.")
            vida = st.number_input("Vida útil (años)", value=st.session_state["form_vida"], help="Duración esperada del dispositivo antes de desecharse o reemplazarse.")

        with colB:
            energia_renovable = st.slider("Energía renovable (%)", 0, 100, st.session_state["form_energia_renovable"], help="Porcentaje de energía que proviene de fuentes renovables.")
            funcionalidad = st.slider("Funcionalidad (1-10)", 1, 10, st.session_state["form_funcionalidad"], help="Nivel de funcionalidad y utilidad que ofrece el dispositivo.")
            reciclabilidad = st.slider("Reciclabilidad (%)", 0, 100, st.session_state["form_reciclabilidad"], help="Porcentaje del dispositivo que puede reciclarse al finalizar su vida útil.")

        with st.expander("Datos de mantenimiento"):
            colM1, colM2 = st.columns(2)
            with colM1:
                B = st.number_input("Baterías vida útil", value=st.session_state["form_B"], help="Cantidad de baterías necesarias durante toda la vida útil del dispositivo.")
                Wb = st.number_input("Peso batería (g)", value=st.session_state["form_Wb"], help="Peso de cada batería en gramos.")
                M = st.number_input("Mantenimientos", value=st.session_state["form_M"], help="Número de veces que el dispositivo requiere mantenimiento.")
                C = st.number_input("Componentes reemplazados", value=st.session_state["form_C"], help="Número de componentes reemplazados en mantenimientos.")

            with colM2:
                Wc = st.number_input("Peso componente (g)", value=st.session_state["form_Wc"], help="Peso promedio de cada componente reemplazado en gramos.")
                W0 = st.number_input("Peso nuevo (g)", value=st.session_state["form_W0"], help="Peso total del dispositivo cuando es nuevo.")
                W = st.number_input("Peso final (g)", value=st.session_state["form_W"], help="Peso final del dispositivo después del uso.")

        submitted = st.form_submit_button("Añadir dispositivo")

with col2:
    st.subheader("Ajuste de Pesos por Métrica")

    with st.expander("Información sobre los métodos de asignación de pesos"):
        st.markdown("""
        ### Método de Asignación de Pesos
        El sistema ofrece tres formas de asignar pesos a las métricas:

        1. **Pesos Recomendados**: 
           - Basados en análisis y alineación con ODS
           - Ideal para usuarios que buscan una evaluación estándar
           - No requiere configuración adicional

        2. **Ajuste Manual**:
           - Permite personalizar los pesos según necesidades específicas
           - Útil cuando se tiene conocimiento experto del dominio
           - Los pesos deben sumar 1.0

        3. **Calcular nuevos pesos**:
           - Utiliza la Matriz de Comparación por Pares
           - Requiere evaluar la importancia relativa entre métricas
           - Incluye verificación de consistencia
        """)

    modo_pesos = st.radio(
        "Selecciona el método de asignación de pesos:",
        ("Pesos Recomendados", "Ajuste Manual", "Calcular nuevos pesos"),
        key="modo_pesos_radio"
    )

    if modo_pesos == "Pesos Recomendados":
        pesos_usuario = obtener_pesos_recomendados()
        st.success("Se han cargado los pesos recomendados del modelo AHP+ODS.")

        df_pesos = pd.DataFrame.from_dict(pesos_usuario, orient='index', columns=['Peso'])
        df_pesos.index = df_pesos.index.map(NOMBRES_METRICAS)
        df_pesos = df_pesos.rename_axis('Métrica').reset_index()
        
        # Agregar columna de importancia relativa
        df_pesos['Importancia'] = df_pesos['Peso'].apply(
            lambda x: '🔴 Alta' if x >= 0.20 else '🟡 Media' if 0.10 < x < 0.20 else '🟢 Baja'
        )
        
        st.dataframe(df_pesos.style.format({'Peso': '{:.3f}'}), use_container_width=True)

        with st.expander("Ver explicación del modelo AHP+ODS"):
            st.markdown("""
            **Metodología de Cálculo de Pesos Recomendados**

            Los pesos recomendados se calcularon combinando dos enfoques complementarios:

            **1. Alineación con los Objetivos de Desarrollo Sostenible (ODS):**
            - Cada métrica fue evaluada según cuántos ODS relevantes aborda
            - Por ejemplo, el consumo de energía se alinea con el ODS 7 (energía asequible y no contaminante) y el ODS 13 (acción por el clima)

            **2. Evaluación cualitativa del impacto ambiental directo:**
            - Se asignó a cada métrica una puntuación de 1 a 5 según:
              - Magnitud del impacto
              - Frecuencia del impacto
              - Alcance del impacto
            - Por ejemplo, la huella de carbono y el consumo de energía obtienen una puntuación más alta que el mantenimiento

            **Cálculo del Score Total:**
            ```
            Score = 1 × ODS + 2 × Impacto
            ```
            Donde:
            - ODS: Número de ODS relevantes que aborda la métrica
            - Impacto: Puntuación cualitativa del impacto ambiental (1-5)

            **Aplicación del Método AHP:**
            - Con base en estos scores, se construyó una matriz de comparación por pares entre métricas
            - Se utilizó la escala de Saaty (1, 3, 5, 7, 9 y sus recíprocos)
            - Esta matriz refleja qué tan importante es una métrica en relación con otra

            **Resultado Final:**
            - Se calculó un vector de pesos normalizado
            - Todos los pesos suman 1
            - Representa la importancia relativa de cada métrica dentro del índice de sostenibilidad ambiental
            """)

    elif modo_pesos == "Ajuste Manual":
        st.info("""
        **Instrucciones para el Ajuste Manual:**
        - Asigne un peso entre 0 y 1 a cada métrica
        - La suma total debe ser 1.0
        - Los pesos más altos indican mayor importancia
        - El sistema normalizará automáticamente si la suma no es 1.0
        """)

        with st.expander("Gestión de configuraciones personalizadas"):
            nuevo_nombre = st.text_input("Guardar configuración como", "")
            guardar = st.button("Guardar configuración")
            if guardar and nuevo_nombre:
                config_actual = {k: float(st.session_state[f"peso_manual_{k}"]) for k in NOMBRES_METRICAS}
                if 'pesos_guardados' not in st.session_state:
                    st.session_state.pesos_guardados = {}
                st.session_state.pesos_guardados[nuevo_nombre] = config_actual
                st.success(f"Configuración '{nuevo_nombre}' guardada correctamente.")
                st.rerun()
            # Mostrar selectbox y botones para aplicar/eliminar si hay configuraciones guardadas
            if st.session_state.get('pesos_guardados'):
                seleccion = st.selectbox("Seleccionar configuración guardada", list(st.session_state.pesos_guardados.keys()))
                col_config = st.columns(2)
                if col_config[0].button("Aplicar configuración"):
                    st.session_state.pesos_manuales = st.session_state.pesos_guardados[seleccion].copy()
                    for k in NOMBRES_METRICAS:
                        st.session_state[f"peso_manual_{k}"] = st.session_state.pesos_manuales[k]
                    st.success(f"Configuración '{seleccion}' aplicada correctamente.")
                    st.rerun()
                if col_config[1].button("Eliminar configuración"):
                    del st.session_state.pesos_guardados[seleccion]
                    st.success(f"Configuración '{seleccion}' eliminada correctamente.")
                    st.rerun()
            if st.button("Reiniciar configuración"):
                inicializar_pesos_manuales()
                st.rerun()

        # Mensaje de configuración activa antes de los inputs manuales
        pesos_actuales = {k: st.session_state.get(f"peso_manual_{k}", 0) for k in NOMBRES_METRICAS}
        pesos_recomendados = obtener_pesos_recomendados()
        if to_dict_flat(pesos_actuales) == to_dict_flat(pesos_recomendados):
            st.success("**Configuración activa: Pesos Recomendados**")
        else:
            st.success("**Configuración activa: Pesos Manuales Personalizados**")

        pesos_usuario = {}
        for id, nombre_metrica in NOMBRES_METRICAS.items():
            valor_default = float(st.session_state.get(f"peso_manual_{id}", obtener_pesos_recomendados()[id]))
            valor = st.number_input(
                f"{nombre_metrica}",
                min_value=0.0,
                max_value=1.0,
                step=0.01,
                format="%.3f",
                value=valor_default,
                key=f"peso_manual_{id}"
            )
            pesos_usuario[id] = float(valor)
        # Asegurar que todas las métricas están presentes
        for id in NOMBRES_METRICAS:
            if id not in pesos_usuario:
                pesos_usuario[id] = float(st.session_state.get(f"peso_manual_{id}", obtener_pesos_recomendados()[id]))
        total_temporal = sum(pesos_usuario.values())
        st.metric("Suma total de pesos", f"{total_temporal:.3f}")
        # No actualizar session_state.pesos_manuales aquí
        # Solo usar pesos_usuario para validación y visualización

        pesos_usuario, es_valido = validar_pesos_manuales(pesos_usuario)
        if not es_valido:
            st.warning("Los pesos fueron normalizados automáticamente para que sumen 1.")
            with st.expander("Ver pesos normalizados"):
                st.write("Los siguientes pesos normalizados serán utilizados en el cálculo:")
                for id, valor in pesos_usuario.items():
                    st.write(f"{NOMBRES_METRICAS[id]}: {valor:.3f}")
                st.write(f"Suma total normalizada: {sum(pesos_usuario.values()):.3f}")

    elif modo_pesos == "Calcular nuevos pesos":
        st.info("""
        **Matriz de Comparación por Pares:**
        - Este método le permite comparar la importancia relativa de cada par de métricas
        - Se utiliza la escala de Saaty para las comparaciones
        - El sistema verificará la consistencia de sus comparaciones
        - Se recomienda comenzar comparando las métricas más importantes entre sí
        """)
        # Mostrar configuración activa si hay pesos calculados
        if 'pesos_ahp' in st.session_state and st.session_state.pesos_ahp is not None:
            nombre_config = "Pesos Calculados"  # valor por defecto
            for nombre_config, config in st.session_state.configuraciones_ahp.items():
                if to_dict_flat(config['pesos']) == to_dict_flat(st.session_state.pesos_ahp):
                    nombre_config = f"Configuración Calculada: {nombre_config}"
                    break
            st.success(f"**Configuración activa:** {nombre_config}")
        if st.button("Editar matriz de comparación por pares"):
            st.session_state.modo_pesos_guardado = st.session_state.modo_pesos_radio
            st.session_state.matriz_ahp_abierta = True
            st.rerun()
        # Mostrar la tabla resumen de pesos calculados si existen
        if 'pesos_ahp' in st.session_state and st.session_state.pesos_ahp is not None:
            # Mostrar solo la tabla de pesos sin el mensaje
            pesos = st.session_state.pesos_ahp
            pesos_limpios = {}
            for k, v in pesos.items():
                if isinstance(v, dict):
                    val = list(v.values())[0]
                elif isinstance(v, pd.Series):
                    val = v.iloc[0]
                else:
                    val = v
                pesos_limpios[k] = float(val)
            # Crear DataFrame con los pesos
            metricas_list = list(pesos_limpios.keys())
            pesos_list = [pesos_limpios[k] for k in metricas_list]
            nombres_list = [NOMBRES_METRICAS[k] for k in metricas_list]
            df_pesos = pd.DataFrame({
                'Métrica': nombres_list,
                'Peso': pesos_list
            })
            # Agregar columna de importancia relativa con criterios equilibrados
            df_pesos['Importancia'] = df_pesos['Peso'].apply(
                lambda x: '🔴 Alta' if x >= 0.20 else '🟡 Media' if 0.10 < x < 0.20 else '🟢 Baja'
            )
            st.dataframe(df_pesos.style.format({'Peso': '{:.3f}'}), use_container_width=True)
            # Añadir expander de gestión de configuraciones de pesos calculados
            with st.expander("Gestión de configuraciones de pesos calculados"):
                if 'configuraciones_ahp' not in st.session_state:
                    st.session_state.configuraciones_ahp = {}
                nuevo_nombre = st.text_input("Guardar configuración como", "")
                if st.button("Guardar configuración") and nuevo_nombre:
                    config_actual = {
                        'pesos': st.session_state.pesos_ahp,
                        'rc': st.session_state.ahp_resultados['rc'] if 'ahp_resultados' in st.session_state else None,
                        'matriz': st.session_state.matriz_comparacion.copy()
                    }
                    st.session_state.configuraciones_ahp[nuevo_nombre] = config_actual
                    st.success(f"Configuración '{nuevo_nombre}' guardada correctamente.")
                    st.rerun()
                if st.session_state.configuraciones_ahp:
                    seleccion = st.selectbox("Seleccionar configuración guardada", list(st.session_state.configuraciones_ahp.keys()))
                    col_config = st.columns(2)
                    if col_config[0].button("Aplicar configuración"):
                        config = st.session_state.configuraciones_ahp[seleccion]
                        st.session_state.pesos_ahp = config['pesos']
                        st.session_state.matriz_comparacion = config['matriz']
                        if 'ahp_resultados' not in st.session_state:
                            st.session_state.ahp_resultados = {}
                        st.session_state.ahp_resultados['pesos'] = config['pesos']
                        st.session_state.ahp_resultados['rc'] = config['rc']
                        st.success(f"Configuración '{seleccion}' aplicada correctamente.")
                        st.rerun()
                    if col_config[1].button("Eliminar configuración"):
                        del st.session_state.configuraciones_ahp[seleccion]
                        st.success(f"Configuración '{seleccion}' eliminada correctamente.")
                        st.rerun()

if submitted:
    # Eliminar solo los resultados globales y la fecha, NO los pesos AHP
    for var in ["resultado_global", "fecha_calculo_global"]:
        if var in st.session_state:
            del st.session_state[var]

    # Determinar los pesos activos y el nombre de la configuración en este momento
    if st.session_state.modo_pesos_radio == "Calcular nuevos pesos":
        if 'pesos_ahp' in st.session_state:
            pesos_usuario = st.session_state.pesos_ahp
            # Buscar nombre de la configuración calculada activa
            nombre_config_pesos = "Pesos Calculados"
            for nombre_config_ahp, config in st.session_state.configuraciones_ahp.items():
                if to_dict_flat(config['pesos']) == to_dict_flat(pesos_usuario):
                    nombre_config_pesos = f"Configuración Calculada: {nombre_config_ahp}"
                    break
        else:
            st.warning("No hay pesos AHP calculados. Se usarán los pesos recomendados.")
            pesos_usuario = obtener_pesos_recomendados()
            nombre_config_pesos = "Pesos Recomendados"
    elif st.session_state.modo_pesos_radio == "Ajuste Manual":
        pesos_manuales = {k: st.session_state[f"peso_manual_{k}"] for k in NOMBRES_METRICAS}
        pesos_usuario, _ = validar_pesos_manuales(pesos_manuales)
        # Buscar nombre de la configuración manual activa
        nombre_config_pesos = "Pesos Manuales Personalizados"
        for nombre_config_manual, config in st.session_state.pesos_guardados.items():
            if to_dict_flat(config) == to_dict_flat(pesos_usuario):
                nombre_config_pesos = f"Configuración Manual: {nombre_config_manual}"
                break
    else:
        pesos_usuario = obtener_pesos_recomendados()
        nombre_config_pesos = "Pesos Recomendados"

    # Calcular el índice de sostenibilidad usando estos pesos
    sensor = SostenibilidadIoT(nombre)
    pesos_limpios = {}
    for k, v in pesos_usuario.items():
        if isinstance(v, dict):
            v = list(v.values())[0]
        try:
            pesos_limpios[k] = float(v)
        except Exception:
            continue
    sensor.pesos = pesos_limpios
    sensor.calcular_consumo_energia(potencia, horas, dias)
    sensor.calcular_huella_carbono()
    sensor.calcular_ewaste(peso, vida)
    sensor.calcular_energia_renovable(energia_renovable)
    sensor.calcular_eficiencia_energetica(funcionalidad)
    sensor.calcular_durabilidad(vida)
    sensor.calcular_reciclabilidad(reciclabilidad)
    sensor.calcular_indice_mantenimiento(B, Wb, M, C, Wc, W0, W)
    resultado = sensor.calcular_sostenibilidad()

    dispositivo_data = {
        "id": str(uuid.uuid4()),
        "nombre": nombre,
        "potencia": potencia,
        "horas": horas,
        "dias": dias,
        "peso": peso,
        "vida": vida,
        "energia_renovable": energia_renovable,
        "funcionalidad": funcionalidad,
        "reciclabilidad": reciclabilidad,
        "B": B,
        "Wb": Wb,
        "M": M,
        "C": C,
        "Wc": Wc,
        "W0": W0,
        "W": W,
        "calculo_realizado": True,
        "pesos_utilizados": pesos_usuario,
        "resultado": resultado,
        # Guardar snapshot del formulario y pesos
        "snapshot_form": {k: st.session_state[f"form_{k}"] for k in form_keys},
        "snapshot_pesos": {
            "modo": st.session_state.modo_pesos_radio,
            "nombre_configuracion": nombre_config_pesos,
            "pesos_manuales": st.session_state.get("pesos_manuales", {}),
            "pesos_ahp": st.session_state.get("pesos_ahp", {})
        }
    }

    st.session_state.dispositivos.append(dispositivo_data)

    st.success(f"Dispositivo '{nombre}' añadido correctamente. Presiona 'Calcular Índice de Sostenibilidad Total' para ver los resultados.")
    st.rerun()

# --- BOTÓN DE REFRESH ---
if 'modo_edicion' in st.session_state and st.session_state.modo_edicion:
    st.warning("Termina de editar o cancelar la edición de un dispositivo antes de calcular el índice global.")
    st.button("Calcular Indice de Sostenibilidad", disabled=True)
else:
    if st.button("Calcular Indice de Sostenibilidad"):
        if not st.session_state.dispositivos:
            st.warning("No hay dispositivos añadidos.")
        else:
            total_indices = []
            metricas_totales = []

            for idx, dispositivo in enumerate(st.session_state.dispositivos):
                # Si no existe resultado, recalcularlo usando los pesos guardados
                if "resultado" not in dispositivo or not dispositivo["calculo_realizado"]:
                    sensor = SostenibilidadIoT(dispositivo["nombre"])
                    sensor.pesos = {k: float(list(v.values())[0]) if isinstance(v, dict) else float(v) for k, v in dispositivo["pesos_utilizados"].items()}
                    sensor.calcular_consumo_energia(dispositivo["potencia"], dispositivo["horas"], dispositivo["dias"])
                    sensor.calcular_huella_carbono()
                    sensor.calcular_ewaste(dispositivo["peso"], dispositivo["vida"])
                    sensor.calcular_energia_renovable(dispositivo["energia_renovable"])
                    sensor.calcular_eficiencia_energetica(dispositivo["funcionalidad"])
                    sensor.calcular_durabilidad(dispositivo["vida"])
                    sensor.calcular_reciclabilidad(dispositivo["reciclabilidad"])
                    sensor.calcular_indice_mantenimiento(
                        dispositivo["B"], dispositivo["Wb"], dispositivo["M"], 
                        dispositivo["C"], dispositivo["Wc"], dispositivo["W0"], dispositivo["W"]
                    )
                    resultado = sensor.calcular_sostenibilidad()
                    dispositivo["resultado"] = resultado
                    dispositivo["calculo_realizado"] = True

                # Actualizar listas para el cálculo global
                total_indices.append(dispositivo["resultado"]["indice_sostenibilidad"])
                metricas_totales.append(dispositivo["resultado"]["metricas_normalizadas"])

            # Cálculo del índice global
            promedio_total = sum(total_indices) / len(total_indices)

            # Calcular promedio de métricas
            promedio_metricas = {
                key: sum(m[key] for m in metricas_totales) / len(metricas_totales)
                for key in metricas_totales[0]
            }

            # Guardar el resultado global
            st.session_state.resultado_global = {
                "promedio_total": promedio_total,
                "promedio_metricas": promedio_metricas
            }

            st.success("Resultados actualizados correctamente.")

# --- FUNCIÓN PARA MOSTRAR RESULTADOS DE UN DISPOSITIVO ---
def mostrar_dispositivo(dispositivo, idx):
    """Muestra el gráfico, recomendaciones y detalles completos del dispositivo."""
    if dispositivo.get("calculo_realizado", False) and "resultado" in dispositivo:
        resultado = dispositivo["resultado"]
        nombre = dispositivo["nombre"]

        # Crear columnas para organizar la disposición
        col1, col2 = st.columns([2, 1])

        # Gráfico a la izquierda
        with col1:
            st.markdown(f"### {nombre}")
            radar_chart(
                resultado["metricas_normalizadas"], 
                f"Métricas Normalizadas - {nombre}",
                key=f"radar_{idx}"
            )

        # Recomendaciones a la derecha
        with col2:
            st.metric("Índice de Sostenibilidad", f"{resultado['indice_sostenibilidad']:.2f}/10")
            st.markdown("### Recomendaciones")
            recomendaciones = []

            if resultado['metricas_normalizadas']['ER'] < 5:
                recomendaciones.append("Aumentar uso de energía renovable.")
            if resultado['metricas_normalizadas']['DP'] < 5:
                recomendaciones.append("Incrementar durabilidad del hardware.")
            if resultado['metricas_normalizadas']['IM'] < 5:
                recomendaciones.append("Reducir impacto de mantenimiento.")
            if resultado['metricas_normalizadas']['EE'] < 5:
                recomendaciones.append("Mejorar eficiencia energética.")
            if resultado['indice_sostenibilidad'] < 6:
                recomendaciones.append("Revisar métricas críticas para mejorar el índice global.")

            for rec_idx, rec in enumerate(recomendaciones):
                st.button(rec, disabled=True, key=f"rec_{idx}_{rec_idx}")

        # --- Secciones de detalles ---
        st.markdown("#### Detalles completos del dispositivo")
        with st.expander("Datos de entrada del dispositivo"):
            nombres_campos = {
                'nombre': 'Nombre del dispositivo',
                'potencia': 'Potencia (W)',
                'horas': 'Horas de uso diario',
                'dias': 'Días de uso al año',
                'peso': 'Peso del dispositivo (kg)',
                'vida': 'Vida útil (años)',
                'energia_renovable': 'Energía renovable (%)',
                'funcionalidad': 'Funcionalidad (1-10)',
                'reciclabilidad': 'Reciclabilidad (%)',
                'B': 'Baterías vida útil',
                'Wb': 'Peso batería (g)',
                'M': 'Mantenimientos',
                'C': 'Componentes reemplazados',
                'Wc': 'Peso componente (g)',
                'W0': 'Peso nuevo (g)',
                'W': 'Peso final (g)'
            }
            datos = {nombres_campos[k]: v for k, v in dispositivo.items() if k in nombres_campos}
            df_datos = pd.DataFrame(datos, index=[0]).T
            df_datos.columns = ['Valor']
            st.dataframe(df_datos, use_container_width=True)

        with st.expander("Pesos utilizados"):
            pesos = dispositivo.get('pesos_utilizados', {})
            # Limpiar y asegurar que todos los valores sean floats
            pesos_limpios = {}
            for k, v in pesos.items():
                if isinstance(v, dict):
                    v = list(v.values())[0]
                try:
                    pesos_limpios[k] = float(v)
                except Exception:
                    continue
            if pesos_limpios:
                # Mostrar el nombre de la configuración guardado explícitamente
                nombre_config = dispositivo.get('snapshot_pesos', {}).get('nombre_configuracion', 'Desconocido')
                st.markdown(f"**Configuración de pesos:** {nombre_config}")
                df_pesos = pd.DataFrame.from_dict(pesos_limpios, orient='index', columns=['Peso'])
                df_pesos.index = df_pesos.index.map(NOMBRES_METRICAS)
                df_pesos = df_pesos.rename_axis('Métrica').reset_index()
                st.dataframe(df_pesos.style.format({'Peso': '{:.3f}'}), use_container_width=True)
            else:
                st.info('No se registraron pesos para este dispositivo.')

        # Checkbox y confirmación para eliminar
        eliminar_key = f'eliminar_{dispositivo["id"]}'
        if st.checkbox('Eliminar dispositivo', key=eliminar_key):
            st.warning('¿Estás seguro de que deseas eliminar este dispositivo? Esta acción no se puede deshacer.')
            if st.button('Confirmar eliminación', key=f'confirmar_{dispositivo["id"]}'):
                st.session_state.dispositivos = [d for d in st.session_state.dispositivos if d["id"] != dispositivo["id"]]
                # Eliminar resultados globales y relacionados siempre que se elimina un dispositivo
                for var in ["resultado_global", "fecha_calculo_global", "mostrar_tabla_pesos_ahp", "pesos_ahp", "ahp_resultados"]:
                    if var in st.session_state:
                        del st.session_state[var]
                st.success(f"Dispositivo '{nombre}' eliminado correctamente.")
                st.rerun()

# --- MOSTRAR RESULTADOS INDIVIDUALES ---
if st.session_state.dispositivos:
    st.markdown("---")
    st.subheader("Resultados por Dispositivo")

    dispositivos = st.session_state.dispositivos
    num_dispositivos = len(dispositivos)

    # Mostrar resultados individuales de todos los dispositivos
    for idx, disp in enumerate(dispositivos):
        # Si no existe resultado, recalcularlo usando los pesos guardados
        if "resultado" not in disp or not disp["calculo_realizado"]:
            sensor = SostenibilidadIoT(disp["nombre"])
            sensor.pesos = {k: float(list(v.values())[0]) if isinstance(v, dict) else float(v) for k, v in disp["pesos_utilizados"].items()}
            sensor.calcular_consumo_energia(disp["potencia"], disp["horas"], disp["dias"])
            sensor.calcular_huella_carbono()
            sensor.calcular_ewaste(disp["peso"], disp["vida"])
            sensor.calcular_energia_renovable(disp["energia_renovable"])
            sensor.calcular_eficiencia_energetica(disp["funcionalidad"])
            sensor.calcular_durabilidad(disp["vida"])
            sensor.calcular_reciclabilidad(disp["reciclabilidad"])
            sensor.calcular_indice_mantenimiento(
                disp["B"], disp["Wb"], disp["M"], disp["C"], disp["Wc"], disp["W0"], disp["W"]
            )
            resultado = sensor.calcular_sostenibilidad()
            disp["resultado"] = resultado
            disp["calculo_realizado"] = True
        # Mostrar resumen y botón de expandir/ocultar detalles
        with st.container():
            col_res, col_btn_det = st.columns([5, 1])
            col_res.markdown(f"**{disp['nombre']}** — Índice: {disp['resultado']['indice_sostenibilidad']:.2f}/10")
            key_exp = f"expandir_disp_{disp['id']}"
            if key_exp not in st.session_state:
                st.session_state[key_exp] = False
            if col_btn_det.button("Mostrar detalles" if not st.session_state[key_exp] else "Ocultar detalles", key=f"btn_toggle_{disp['id']}"):
                st.session_state[key_exp] = not st.session_state[key_exp]
                st.rerun()
            if st.session_state[key_exp]:
                mostrar_dispositivo(disp, disp['id'])

# --- MOSTRAR RESULTADOS GLOBALES ---
if "resultado_global" in st.session_state:
    st.markdown("---")
    resultado_global = st.session_state.resultado_global
    promedio_total = resultado_global["promedio_total"]
    promedio_metricas = resultado_global["promedio_metricas"]

    # Gráfico del promedio a la izquierda
    col1, col2 = st.columns([2, 1])
    with col1:
        radar_chart(promedio_metricas, "Promedio de Métricas Normalizadas", key="radar_total")
    # Recomendaciones globales a la derecha
    with col2:
        st.metric("Índice de Sostenibilidad Global", f"{promedio_total:.2f}/10")
        st.markdown("### Recomendaciones Globales")
        recomendaciones_globales = []
        if promedio_metricas["ER"] < 5:
            recomendaciones_globales.append("Aumentar uso de energía renovable.")
        if promedio_metricas["DP"] < 5:
            recomendaciones_globales.append("Incrementar durabilidad del hardware.")
        if promedio_metricas["IM"] < 5:
            recomendaciones_globales.append("Reducir impacto de mantenimiento.")
        if promedio_metricas["EE"] < 5:
            recomendaciones_globales.append("Mejorar eficiencia energética.")
        if promedio_total < 6:
            recomendaciones_globales.append("Revisar métricas críticas para mejorar el índice global.")
        for rec_idx, rec in enumerate(recomendaciones_globales):
            st.button(rec, disabled=True, key=f"global_rec_{rec_idx}")

    # --- Detalles del sistema ---
    with st.expander("Detalles del sistema"):
        dispositivos = st.session_state.dispositivos
        st.markdown(f"**Cantidad total de dispositivos evaluados:** {len(dispositivos)}")
        indices = [d['resultado']['indice_sostenibilidad'] for d in dispositivos if 'resultado' in d]
        if len(indices) > 1:
            std = np.std(indices)
            st.markdown(f"**Desviación estándar de los índices individuales:** {std:.2f}")
        else:
            st.markdown("**Desviación estándar de los índices individuales:** N/A")
        if 'fecha_calculo_global' not in st.session_state:
            st.session_state.fecha_calculo_global = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        st.markdown(f"**Fecha y hora del cálculo global:** {st.session_state.fecha_calculo_global}")
        pesos_usados = [str(d.get('pesos_utilizados', {})) for d in dispositivos]
        if len(set(pesos_usados)) > 1:
            st.warning("Atención: los dispositivos fueron evaluados con diferentes configuraciones de pesos. Los índices individuales pueden no ser directamente comparables.")
        st.markdown("**Pesos utilizados para el cálculo global**")
        if "pesos_ahp" in st.session_state and st.session_state.modo_pesos_radio == "Calcular nuevos pesos":
            pesos_global = st.session_state.pesos_ahp
            nombre_config = "Pesos Calculados"
            for nombre_config, config in st.session_state.configuraciones_ahp.items():
                if to_dict_flat(config['pesos']) == to_dict_flat(pesos_global):
                    nombre_config = f"Configuración Calculada: {nombre_config}"
                    break
        elif st.session_state.modo_pesos_radio == "Ajuste Manual":
            pesos_manuales = {k: st.session_state[f"peso_manual_{k}"] for k in NOMBRES_METRICAS}
            pesos_global, _ = validar_pesos_manuales(pesos_manuales)
            nombre_config = "Pesos Manuales Personalizados"
            for nombre_config, config in st.session_state.pesos_guardados.items():
                if to_dict_flat(config) == to_dict_flat(pesos_global):
                    nombre_config = f"Configuración Manual: {nombre_config}"
                    break
        else:
            pesos_global = obtener_pesos_recomendados()
            nombre_config = "Pesos Recomendados"
        st.markdown(f"**Configuración de pesos:** {nombre_config}")
        pesos_limpios = {}
        for k, v in pesos_global.items():
            if isinstance(v, dict):
                v = list(v.values())[0]
            try:
                pesos_limpios[k] = float(v)
            except Exception:
                continue
        df_pesos = pd.DataFrame.from_dict(pesos_limpios, orient='index', columns=['Peso'])
        df_pesos.index = df_pesos.index.map(NOMBRES_METRICAS)
        df_pesos = df_pesos.rename_axis('Métrica').reset_index()
        st.dataframe(df_pesos.style.format({'Peso': '{:.3f}'}), use_container_width=True)
        st.markdown("**Dispositivos incluidos en el cálculo global**")
        if dispositivos:
            data_disp = {
                'Nombre': [d['nombre'] for d in dispositivos],
                'Índice de Sostenibilidad': [d['resultado']['indice_sostenibilidad'] if 'resultado' in d else None for d in dispositivos]
            }
            df_disp = pd.DataFrame(data_disp)
            st.dataframe(df_disp.style.format({'Índice de Sostenibilidad': '{:.2f}'}), use_container_width=True)
        else:
            st.info('No hay dispositivos incluidos actualmente.')

    # Botón de descarga directo
    excel_file = exportar_resultados_excel()
    st.download_button(
        label="Descargar Resultados Completos",
        data=excel_file,
        file_name=f"sostenibilidad_iot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
